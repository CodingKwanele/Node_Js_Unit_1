import json
import openpyxl
import argparse
from pathlib import Path

def xlsx_to_json(input_file, output_file=None, sheet_name=None, indent=4):
    """
    Convert an Excel (.xlsx) file to JSON format.
    Always uses row 1 as the header row.
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str, optional): Path to the output JSON file. If None, uses input file name with .json extension
        sheet_name (str, optional): Specific sheet to convert. If None, uses the first sheet
        indent (int, optional): JSON indentation level. Set to None for compact JSON
    """
    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return False
    
    # Select the worksheet
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            return False
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
    
    # Get headers from row 1 (first row)
    headers = [2]
    if worksheet.max_row < 1:
        print("The worksheet is empty.")
        return False
    for cell in worksheet[2]:  # Explicitly using row 1
        value = cell.value
        # Ensure header is a string and not empty
        header = str(value) if value is not None else f"column{cell.column}"
        headers.append(header)
    
    # Process data rows (starting from row 2)
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            # Convert None to empty string if needed
            if value is None:
                value = ""
            row_data[header] = value
        data.append(row_data)
    
    # Determine output file path
    if output_file is None:
        input_path = Path(input_file)
        output_file = input_path.with_suffix('.json')
    
    # Write JSON to file
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
        print(f"Successfully converted '{input_file}' to '{output_file}'")
        print(f"Used row 1 as headers: {headers}")
        return True
    except Exception as e:
        print(f"Error writing JSON file: {e}")
        return False

if __name__ == "__main__":
    # Set up command line arguments
    parser = argparse.ArgumentParser(description='Convert XLSX file to JSON (using row 1 as headers)')
    parser.add_argument('input_file', help='Path to the input Excel (.xlsx) file')
    parser.add_argument('-o', '--output', help='Path to the output JSON file (optional)')
    parser.add_argument('-s', '--sheet', help='Specific sheet name to convert (optional)')
    parser.add_argument('-i', '--indent', type=int, default=4, 
                        help='Indentation level for JSON output (default: 4)')
    
    args = parser.parse_args()
    
    # Perform conversion
    xlsx_to_json(
        input_file=args.input_file,
        output_file=args.output,
        sheet_name=args.sheet,
        indent=args.indent
    )